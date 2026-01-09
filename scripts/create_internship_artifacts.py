from openpyxl import Workbook
import os

out_dir = os.path.join(os.getcwd(), 'Internship_artifacts', 'Internship_artifacts')
os.makedirs(out_dir, exist_ok=True)

# Agile Template
wb = Workbook()
ws = wb.active
ws.title = 'Backlog'
headers = ['Task ID', 'Title', 'Description', 'Owner', 'Priority', 'Status', 'Sprint', 'Start Date', 'End Date', 'Comments']
ws.append(headers)
# project-specific backlog items
rows = [
	['T-001', 'Frontend scaffold (Vite + React)', 'Create Vite project, React Router, initial pages', 'Preeti', 'High', 'Done', 'Sprint 1', '2025-12-10', '2025-12-15', ''],
	['T-002', 'Search API integration', 'Integrate /search endpoint, filters and results', 'Preeti', 'High', 'Done', 'Sprint 1', '2025-12-16', '2025-12-20', ''],
	['T-003', 'Flight Details & Price Polling', 'Show flight details and poll dynamic price endpoint', 'Preeti', 'High', 'In Progress', 'Sprint 2', '2025-12-21', '2026-01-05', 'Polling every 10s'],
	['T-004', 'Seat selector UI', 'Interactive seat map with class-based pricing', 'Preeti', 'Medium', 'To Do', 'Sprint 3', '', '', 'Requires seat metadata from backend'],
	['T-005', 'Booking hold & confirm', 'Implement hold (initiate) and confirm endpoints in frontend flow', 'Preeti', 'High', 'Done', 'Sprint 2', '2026-01-03', '2026-01-06', 'Uses sessionStorage to persist hold id'],
	['T-006', 'Passenger multi-entry & validation', 'Allow 1-9 passengers; inline validation for email/name', 'Preeti', 'High', 'Done', 'Sprint 3', '2026-01-06', '2026-01-08', 'Client-side validation implemented'],
	['T-007', 'Auth: Register/Login', 'Add register and login pages, session restore via /me', 'Preeti', 'High', 'Done', 'Sprint 2', '2025-12-25', '2026-01-02', 'Credentials: cookies included'],
	['T-008', 'Booking history & receipts', 'List user bookings and allow JSON/PDF download', 'Preeti', 'Medium', 'Done', 'Sprint 3', '2026-01-07', '2026-01-08', 'PDF endpoint: /bookings/<pnr>/receipt?format=pdf'],
	['T-009', 'Tailwind styling and responsive layout', 'Integrate Tailwind via PostCSS and apply base styles', 'Preeti', 'Medium', 'Done', 'Sprint 1', '2025-12-14', '2025-12-20', 'Imported tailwind.css in main.jsx'],
	['T-010', 'Tests & Docs', 'Add README, setup guide, unit test plan and presentation', 'Preeti', 'Medium', 'In Progress', 'Sprint 4', '2026-01-08', '', 'Generate docs under docs/'],
	['T-011', 'Concurrency hardening', 'Ensure seat holds expire and confirm is transactional', 'Preeti', 'High', 'To Do', 'Sprint 4', '', '', 'Backend DB transactional logic'],
	['T-012', 'CI / Deployment', 'Add GitHub Actions workflow for tests and deploy', 'Preeti', 'Low', 'To Do', 'Sprint 5', '', '', 'Optional'],
]
for r in rows:
	ws.append(r)
wb.save(os.path.join(out_dir, 'Agile_Template_v0.1.xlsx'))

# Defect Tracker
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'Defects'
headers2 = ['Defect ID', 'Summary', 'Steps to Reproduce', 'Severity', 'Priority', 'Status', 'Assigned To', 'Reported By', 'Date Found', 'Environment', 'Resolution']
ws2.append(headers2)
# project-specific defect examples
defects = [
	['D-001', 'Login returns 500 on invalid payload', '1. POST /login with missing fields\\n2. Observe server 500', 'Major', 'High', 'Open', 'Backend Team', 'QA', '2026-01-03', 'Local', 'Investigate payload handling and input validation'],
	['D-002', 'Seat hold not released on timeout', '1. Initiate hold\\n2. Do not confirm\\n3. Seats stay reserved', 'Critical', 'High', 'Open', 'Backend Team', 'QA', '2026-01-04', 'Local', 'Expiry job not firing'],
	['D-003', 'Price polling causes UI jitter', '1. Open flight details\\n2. Observe layout shift when price updates', 'Minor', 'Low', 'Open', 'Frontend Team', 'QA', '2026-01-06', 'Chrome', 'Stabilize layout and animate price changes'],
	['D-004', 'Register allows duplicate email', '1. Register with existing email\\n2. Account created/exception', 'Major', 'High', 'Open', 'Backend Team', 'QA', '2026-01-07', 'Local', 'Add unique constraint and error handling'],
]
for d in defects:
	ws2.append(d)
wb2.save(os.path.join(out_dir, 'Defect_Tracker_Template_v0.1.xlsx'))

# Unit Test Plan
wb3 = Workbook()
ws3 = wb3.active
ws3.title = 'Unit Tests'
headers3 = ['Test Case ID', 'Module', 'Test Objective', 'Preconditions', 'Test Steps', 'Expected Result', 'Actual Result', 'Status', 'Tester', 'Date']
# project-specific unit test cases
tests = [
	['UT-001', 'backend:/search', 'Verify /search returns results for valid query', 'Backend running, sample data loaded', 'Call GET /search?origin=SFO&destination=LAX', '200 OK and non-empty flights list', '', 'Not Run', 'Preeti', '2026-01-04'],
	['UT-002', 'backend:/flights/<id>/price', 'Verify price endpoint returns numeric dynamic price', 'Backend running', 'Call GET /flights/1/price', '200 OK with price and timestamp', '', 'Not Run', 'Preeti', '2026-01-04'],
	['UT-003', 'backend:/book/initiate', 'Verify initiate booking creates a temporary hold', 'Backend running, seat available', 'POST /book/initiate with flight_id & seats', '200 OK with temp_reference and hold expiry', '', 'Not Run', 'Preeti', '2026-01-05'],
	['UT-004', 'backend:/book/confirm', 'Confirm booking completes and generates PNR', 'Hold exists and valid', 'POST /book/confirm with temp_reference and passenger data', '200 OK with PNR and booking details', '', 'Not Run', 'Preeti', '2026-01-05'],
	['UT-005', 'auth:/register & /login', 'Register creates user and login sets session', 'Backend running', 'POST /register then POST /login', '200 OK for both and session cookie present', '', 'Not Run', 'Preeti', '2026-01-06'],
	['UT-006', 'concurrency:double-book', 'Ensure two simultaneous confirms cannot overbook', 'Simulate two clients holding same seats', 'Simultaneously POST /book/confirm for both', 'Only one should succeed; other returns error', '', 'Not Run', 'Preeti', '2026-01-07'],
	['UT-007', 'frontend:PassengerForm validation', 'Verify inline validation works for name/email', 'Frontend running', 'Interact with PassengerInfo form, leave fields blank or invalid', 'Inline errors shown and Continue blocked', '', 'Not Run', 'Preeti', '2026-01-08'],
	['UT-008', 'frontend:Login/Register flows', 'Verify register and login update UI and session restore', 'Frontend + backend running', 'Register new account, reload, check /me', 'User info returned and UI shows logged-in state', '', 'Not Run', 'Preeti', '2026-01-08'],
	['UT-009', 'receipt:download', 'Verify booking receipt endpoints return JSON and PDF', 'Booking exists', 'GET /bookings/<pnr>/receipt?format=json and ?format=pdf', '200 with appropriate Content-Type and payload', '', 'Not Run', 'Preeti', '2026-01-09'],
	['UT-010', 'ui:price-polling', 'Verify price polling updates without crash', 'Frontend running', 'Open FlightDetails and wait for multiple poll cycles', 'UI shows updated price values and no errors', '', 'Not Run', 'Preeti', '2026-01-09'],
]
for t in tests:
	ws3.append(t)

# Additional user-provided test cases (TC001 - TC005)
user_tests = [
	['TC001', 'Data Preprocessing', 'Verify dataset loads and preprocesses without missing or invalid values', 'Cleaned dataset with no missing or invalid data', 'All columns validated and cleaned', 'Pass', 'Preeti', '2026-01-09', 'Manual'],
	['TC002', 'Model Training', 'Validate trained model predicts fraud correctly on test data', 'Fraud prediction accuracy >= 95%', 'Accuracy achieved 98%', 'Pass', 'Preeti', '2026-01-09', 'Automated'],
	['TC003', 'API Endpoint /upload', 'Ensure CSV upload API accepts file and returns fraud prediction response', 'Returns JSON with predictions for all rows', 'API response validated successfully', 'Pass', 'Preeti', '2026-01-09', 'Integration'],
	['TC004', 'Frontend Upload Component', 'Check React upload component correctly sends file to backend', 'Upload success message shown and data reloads', 'Upload working and UI updates instantly', 'Pass', 'Preeti', '2026-01-09', 'Manual'],
	['TC005', 'Dashboard Analytics', 'Confirm analytics dashboard updates dynamically after upload', 'Charts and transaction table display new results', 'Dashboard refreshed and analytics updated', 'Pass', 'Preeti', '2026-01-09', 'Manual'],
]

# Ensure headers match new rows: append these to the same sheet
for ut in user_tests:
	ws3.append(ut)
wb3.save(os.path.join(out_dir, 'Unit_Test_Plan_v0.1.updated.xlsx'))

print('Created files in', out_dir)
